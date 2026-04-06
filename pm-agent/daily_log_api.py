"""
Daily Log API — 매일의 장부 + 할 일 관리

철학:
- 대표가 매일 아침/저녁 5초만에 매출/주문/광고비 입력
- 체크리스트로 할 일 관리
- 입력된 데이터는 자동으로 KPI 월간 실적에 누적
"""
import json
import logging
import sqlite3
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/daily", tags=["Daily-Log"])
DB_PATH = str(Path(__file__).parent / "data" / "approval_queue.db")


def _init():
    with sqlite3.connect(DB_PATH) as conn:
        # 일일 장부 — 대표의 기존 엑셀 장부 구조 반영
        conn.execute("""
            CREATE TABLE IF NOT EXISTS daily_log (
                date TEXT PRIMARY KEY,
                weekday TEXT DEFAULT '',

                -- 결제/정산 (엑셀 '매출현황' 기준)
                payment_krw REAL DEFAULT 0,              -- 결제금액
                payment_shipping_krw REAL DEFAULT 0,     -- 결제배송비
                settlement_krw REAL DEFAULT 0,           -- 정산예정금액 (플랫폼 수수료 차감)

                -- 원가 (해외 구매 + 배송)
                purchase_cost_krw REAL DEFAULT 0,        -- 구매금액(원화)
                international_shipping_krw REAL DEFAULT 0, -- 국제배송비
                cargo_shipping_krw REAL DEFAULT 0,       -- 화물택배비

                -- 수익 (자동 계산)
                profit_krw REAL DEFAULT 0,               -- 수익금
                profit_rate REAL DEFAULT 0,              -- 수익률

                -- 건수
                orders_count INTEGER DEFAULT 0,          -- 결제건수
                returns_count INTEGER DEFAULT 0,         -- 반품
                cs_count INTEGER DEFAULT 0,              -- CS

                -- 광고/부가
                ad_spend_krw REAL DEFAULT 0,             -- 광고비
                ad_roas REAL DEFAULT 0,                  -- ROAS

                -- 환율 스냅샷
                fx_usd REAL DEFAULT 0,
                fx_cny REAL DEFAULT 0,
                fx_jpy REAL DEFAULT 0,

                -- 메모
                note TEXT DEFAULT '',
                mood TEXT DEFAULT '',
                updated_at TEXT DEFAULT '',

                -- 하위 호환 (이전 필드)
                revenue_krw REAL DEFAULT 0,
                cost_krw REAL DEFAULT 0
            )
        """)

        # 기존 테이블에 신규 컬럼 추가 (ALTER)
        existing_cols = {row[1] for row in conn.execute("PRAGMA table_info(daily_log)").fetchall()}
        new_cols = [
            ("weekday", "TEXT DEFAULT ''"),
            ("payment_krw", "REAL DEFAULT 0"),
            ("payment_shipping_krw", "REAL DEFAULT 0"),
            ("settlement_krw", "REAL DEFAULT 0"),
            ("purchase_cost_krw", "REAL DEFAULT 0"),
            ("international_shipping_krw", "REAL DEFAULT 0"),
            ("cargo_shipping_krw", "REAL DEFAULT 0"),
            ("profit_rate", "REAL DEFAULT 0"),
            ("ad_roas", "REAL DEFAULT 0"),
            ("fx_usd", "REAL DEFAULT 0"),
            ("fx_cny", "REAL DEFAULT 0"),
            ("fx_jpy", "REAL DEFAULT 0"),
        ]
        for col, defn in new_cols:
            if col not in existing_cols:
                try:
                    conn.execute(f"ALTER TABLE daily_log ADD COLUMN {col} {defn}")
                except Exception:
                    pass

        # 주문 원장 — 엑셀 '주문현황' 시트 1건당 1row
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_ledger (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_date TEXT,
                platform TEXT,
                order_no TEXT,
                product_no TEXT,
                product_name TEXT,
                option_name TEXT,
                quantity INTEGER DEFAULT 1,
                product_url TEXT,
                buyer_name TEXT,
                buyer_phone TEXT,
                recipient_name TEXT,
                recipient_phone TEXT,
                customs_no TEXT,
                zipcode TEXT,
                address TEXT,
                delivery_memo TEXT,
                payment_krw REAL DEFAULT 0,
                shipping_krw REAL DEFAULT 0,
                settlement_krw REAL DEFAULT 0,
                order_placed_date TEXT,
                source_vendor TEXT,
                source_order_no TEXT,
                source_tracking TEXT,
                payment_card TEXT,
                foreign_currency TEXT,
                foreign_amount REAL DEFAULT 0,
                purchase_usd REAL DEFAULT 0,
                purchase_krw REAL DEFAULT 0,
                forwarding_center TEXT,
                domestic_carrier TEXT,
                domestic_tracking TEXT,
                international_shipping_krw REAL DEFAULT 0,
                cargo_carrier TEXT,
                cargo_tracking TEXT,
                cargo_shipping_krw REAL DEFAULT 0,
                customs_duty_krw REAL DEFAULT 0,
                profit_krw REAL DEFAULT 0,
                profit_rate REAL DEFAULT 0,
                note TEXT DEFAULT '',
                imported_at TEXT DEFAULT '',
                UNIQUE(order_no, platform)
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_order_date ON order_ledger(order_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_order_platform ON order_ledger(platform)")

        # 세무 증빙 필드 (ALTER) — 과세구분/수수료/부가세/증빙
        ol_cols = {row[1] for row in conn.execute("PRAGMA table_info(order_ledger)").fetchall()}
        tax_cols = [
            ("tax_category", "TEXT DEFAULT ''"),        # 과세/면세/영세/대행수수료
            ("platform_fee_krw", "REAL DEFAULT 0"),     # 플랫폼 수수료 (쿠팡11%/스마트4%)
            ("pg_fee_krw", "REAL DEFAULT 0"),           # PG 수수료
            ("vat_krw", "REAL DEFAULT 0"),              # 부가세 (매출의 10/110)
            ("receipt_type", "TEXT DEFAULT ''"),        # 현금영수증/세금계산서/카드매출전표
            ("receipt_no", "TEXT DEFAULT ''"),          # 발행번호
            ("card_last4", "TEXT DEFAULT ''"),          # 사업용 결제카드 끝 4자리
            ("bank_tx_id", "TEXT DEFAULT ''"),          # 은행 거래내역 매칭
            ("evidence_files", "TEXT DEFAULT ''"),      # JSON: [{type,path},...]
            ("tax_year_month", "TEXT DEFAULT ''"),      # 2025-05 (인덱스 용이)
            ("tax_locked", "INTEGER DEFAULT 0"),        # 신고완료 잠금
        ]
        for col, defn in tax_cols:
            if col not in ol_cols:
                try:
                    conn.execute(f"ALTER TABLE order_ledger ADD COLUMN {col} {defn}")
                except Exception:
                    pass
        conn.execute("CREATE INDEX IF NOT EXISTS idx_order_tax_ym ON order_ledger(tax_year_month)")

        # 기존 행에 tax_year_month 채우기 (order_date에서 추출)
        conn.execute("""
            UPDATE order_ledger
            SET tax_year_month = substr(order_date, 1, 7)
            WHERE (tax_year_month IS NULL OR tax_year_month = '')
              AND order_date IS NOT NULL AND length(order_date) >= 7
        """)
        # 기존 행에 플랫폼 수수료 추정치 자동 계산 (payment_krw * 요율)
        conn.execute("""
            UPDATE order_ledger
            SET platform_fee_krw = ROUND(payment_krw * 0.11, 0)
            WHERE platform_fee_krw = 0 AND platform LIKE '%쿠팡%' AND payment_krw > 0
        """)
        conn.execute("""
            UPDATE order_ledger
            SET platform_fee_krw = ROUND(payment_krw * 0.04, 0)
            WHERE platform_fee_krw = 0 AND (platform LIKE '%스마트%' OR platform LIKE '%네이버%') AND payment_krw > 0
        """)

        # CS 템플릿 (엑셀 'CS 템플릿' 시트)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cs_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL UNIQUE,
                template TEXT NOT NULL,
                source TEXT DEFAULT 'user',
                created_at TEXT DEFAULT ''
            )
        """)
        # 일일 할 일
        conn.execute("""
            CREATE TABLE IF NOT EXISTS daily_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                title TEXT NOT NULL,
                priority INTEGER DEFAULT 1,
                status TEXT DEFAULT 'pending',
                rollover_from TEXT DEFAULT '',
                completed_at TEXT DEFAULT '',
                created_at TEXT DEFAULT ''
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_tasks_date ON daily_tasks(date)")
        conn.commit()

_init()


# ═══════════════════════════════════════════════════════════
# 모델
# ═══════════════════════════════════════════════════════════

class DailyLogRequest(BaseModel):
    date: str  # "2026-04-06"
    revenue_krw: float = 0
    orders_count: int = 0
    ad_spend_krw: float = 0
    cost_krw: float = 0
    profit_krw: float = 0
    returns_count: int = 0
    cs_count: int = 0
    note: str = ""
    mood: str = ""


class TaskRequest(BaseModel):
    date: str
    title: str
    priority: int = 1


class TaskUpdateRequest(BaseModel):
    status: Optional[str] = None
    title: Optional[str] = None
    priority: Optional[int] = None


# ═══════════════════════════════════════════════════════════
# 일일 장부
# ═══════════════════════════════════════════════════════════

@router.get("/log/{target_date}")
async def get_log(target_date: str):
    """특정 날짜 장부 조회"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM daily_log WHERE date=?", (target_date,)).fetchone()
        if row:
            return dict(row)
        return {
            "date": target_date,
            "revenue_krw": 0, "orders_count": 0, "ad_spend_krw": 0,
            "cost_krw": 0, "profit_krw": 0, "returns_count": 0,
            "cs_count": 0, "note": "", "mood": "", "updated_at": ""
        }


@router.post("/log")
async def save_log(request: DailyLogRequest):
    """일일 장부 저장/수정"""
    now = datetime.now().isoformat()
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO daily_log (date, revenue_krw, orders_count, ad_spend_krw, cost_krw,
                                   profit_krw, returns_count, cs_count, note, mood, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(date) DO UPDATE SET
                revenue_krw=excluded.revenue_krw,
                orders_count=excluded.orders_count,
                ad_spend_krw=excluded.ad_spend_krw,
                cost_krw=excluded.cost_krw,
                profit_krw=excluded.profit_krw,
                returns_count=excluded.returns_count,
                cs_count=excluded.cs_count,
                note=excluded.note,
                mood=excluded.mood,
                updated_at=excluded.updated_at
        """, (request.date, request.revenue_krw, request.orders_count, request.ad_spend_krw,
              request.cost_krw, request.profit_krw, request.returns_count, request.cs_count,
              request.note, request.mood, now))
        conn.commit()
    return {"status": "success", "date": request.date}


@router.get("/log-range")
async def get_log_range(start: str, end: str):
    """기간 조회 (7일/30일 그래프용)"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM daily_log WHERE date >= ? AND date <= ? ORDER BY date",
            (start, end)
        ).fetchall()

    # 누락 날짜 0으로 채움
    from datetime import datetime as _dt, timedelta as _td
    start_d = _dt.strptime(start, "%Y-%m-%d").date()
    end_d = _dt.strptime(end, "%Y-%m-%d").date()

    existing = {r["date"]: dict(r) for r in rows}
    result = []
    cur = start_d
    while cur <= end_d:
        key = cur.isoformat()
        if key in existing:
            result.append(existing[key])
        else:
            result.append({
                "date": key, "revenue_krw": 0, "orders_count": 0, "ad_spend_krw": 0,
                "cost_krw": 0, "profit_krw": 0, "returns_count": 0, "cs_count": 0,
                "note": "", "mood": "", "updated_at": ""
            })
        cur += _td(days=1)

    # 합계
    total = {
        "revenue_krw": sum(r["revenue_krw"] or 0 for r in result),
        "orders_count": sum(r["orders_count"] or 0 for r in result),
        "ad_spend_krw": sum(r["ad_spend_krw"] or 0 for r in result),
        "profit_krw": sum(r["profit_krw"] or 0 for r in result),
        "returns_count": sum(r["returns_count"] or 0 for r in result),
        "cs_count": sum(r["cs_count"] or 0 for r in result),
    }

    # 평균
    days = len(result)
    avg = {k: round(v / days, 0) if days > 0 else 0 for k, v in total.items()}

    return {
        "start": start,
        "end": end,
        "days": days,
        "logs": result,
        "total": total,
        "average": avg,
    }


@router.get("/summary")
async def daily_summary():
    """홈 화면 요약 — 오늘/이번 주/이번 달 + 목표 비교"""
    today = date.today()
    yesterday = (today - timedelta(days=1)).isoformat()
    week_start = (today - timedelta(days=today.weekday())).isoformat()
    month_start = today.replace(day=1).isoformat()
    year_start = today.replace(month=1, day=1).isoformat()
    today_str = today.isoformat()
    month_key = today.strftime("%Y-%m")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row

        def _sum(start, end):
            r = conn.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN payment_krw > 0 THEN payment_krw ELSE revenue_krw END), 0) as revenue,
                    COALESCE(SUM(orders_count), 0) as orders,
                    COALESCE(SUM(ad_spend_krw), 0) as ad_spend,
                    COALESCE(SUM(profit_krw), 0) as profit,
                    COALESCE(SUM(settlement_krw), 0) as settlement,
                    COALESCE(SUM(purchase_cost_krw), 0) as purchase_cost,
                    COALESCE(SUM(international_shipping_krw), 0) as intl_shipping,
                    COALESCE(SUM(cargo_shipping_krw), 0) as cargo_shipping
                FROM daily_log WHERE date >= ? AND date <= ?
            """, (start, end)).fetchone()
            return dict(r) if r else {"revenue": 0, "orders": 0, "ad_spend": 0, "profit": 0, "settlement": 0, "purchase_cost": 0, "intl_shipping": 0, "cargo_shipping": 0}

        today_log = _sum(today_str, today_str)
        yesterday_log = _sum(yesterday, yesterday)
        week_log = _sum(week_start, today_str)
        month_log = _sum(month_start, today_str)
        year_log = _sum(year_start, today_str)

        # 이번 달 일수
        import calendar
        month_days = calendar.monthrange(today.year, today.month)[1]
        days_elapsed = today.day
        days_remaining = month_days - days_elapsed

        # 일평균 (이번 달)
        avg_daily_revenue = month_log["revenue"] / days_elapsed if days_elapsed > 0 else 0
        projected_month = avg_daily_revenue * month_days

        # 오늘 할 일 카운트
        today_tasks = conn.execute(
            "SELECT COUNT(*) as c, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END), 0) as done FROM daily_tasks WHERE date=?",
            (today_str,)
        ).fetchone()

        # 전일 대비 증감
        rev_change = 0
        if yesterday_log["revenue"] > 0:
            rev_change = round((today_log["revenue"] - yesterday_log["revenue"]) / yesterday_log["revenue"] * 100, 1)

        # ── 목표 조회 (KPI 테이블) ──
        month_target_row = conn.execute(
            "SELECT target_value FROM kpi_targets WHERE month=? AND kpi_key='revenue_krw'",
            (month_key,)
        ).fetchone()
        month_target = month_target_row["target_value"] if month_target_row else 0

        # 연 목표 — 12개월 KPI 목표 합
        year_targets = conn.execute(
            "SELECT SUM(target_value) as total FROM kpi_targets WHERE month LIKE ? AND kpi_key='revenue_krw'",
            (f"{today.year}-%",)
        ).fetchone()
        year_target = year_targets["total"] if year_targets and year_targets["total"] else 0

        # 하루 목표 = 월 목표 / 월 총 일수
        daily_target = round(month_target / month_days) if month_target > 0 else 0

        # 달성률 계산
        today_pct = round(today_log["revenue"] / daily_target * 100, 1) if daily_target > 0 else 0
        month_pct = round(month_log["revenue"] / month_target * 100, 1) if month_target > 0 else 0
        year_pct = round(year_log["revenue"] / year_target * 100, 1) if year_target > 0 else 0

        # 이번 달 페이스 — 지금까지 페이스 대비 목표 달성 가능 여부
        pace_target = daily_target * days_elapsed  # 지금까지 도달했어야 할 매출
        pace_diff = month_log["revenue"] - pace_target  # 양수 = 페이스 초과, 음수 = 부족

    return {
        "today": {
            "date": today_str,
            "revenue": today_log["revenue"],
            "orders": today_log["orders"],
            "ad_spend": today_log["ad_spend"],
            "profit": today_log["profit"],
            "revenue_change_pct": rev_change,
            "target": daily_target,
            "target_pct": today_pct,
            "target_remaining": max(0, daily_target - today_log["revenue"]),
        },
        "yesterday": yesterday_log,
        "this_week": {
            "start": week_start,
            "revenue": week_log["revenue"],
            "orders": week_log["orders"],
            "profit": week_log["profit"],
        },
        "this_month": {
            "start": month_start,
            "revenue": month_log["revenue"],
            "orders": month_log["orders"],
            "ad_spend": month_log["ad_spend"],
            "profit": month_log["profit"],
            "days_elapsed": days_elapsed,
            "days_remaining": days_remaining,
            "avg_daily_revenue": round(avg_daily_revenue, 0),
            "projected_month_revenue": round(projected_month, 0),
            "target": month_target,
            "target_pct": month_pct,
            "target_remaining": max(0, month_target - month_log["revenue"]),
            "pace_target": pace_target,
            "pace_diff": pace_diff,
        },
        "this_year": {
            "start": year_start,
            "revenue": year_log["revenue"],
            "orders": year_log["orders"],
            "profit": year_log["profit"],
            "target": year_target,
            "target_pct": year_pct,
            "target_remaining": max(0, year_target - year_log["revenue"]),
        },
        "today_tasks": {
            "total": today_tasks["c"] if today_tasks else 0,
            "completed": today_tasks["done"] if today_tasks else 0,
        }
    }


# ═══════════════════════════════════════════════════════════
# 일일 할 일
# ═══════════════════════════════════════════════════════════

@router.get("/tasks")
async def list_tasks(target_date: Optional[str] = None, include_rollover: bool = True):
    """오늘 할 일 + 어제 미완료 자동 이월"""
    if not target_date:
        target_date = date.today().isoformat()
    yesterday = (date.fromisoformat(target_date) - timedelta(days=1)).isoformat()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row

        # 오늘 할 일
        today_tasks = [dict(r) for r in conn.execute(
            "SELECT * FROM daily_tasks WHERE date=? ORDER BY status, priority DESC, id",
            (target_date,)
        ).fetchall()]

        # 어제 미완료 이월
        rollover_tasks = []
        if include_rollover:
            rollover_tasks = [dict(r) for r in conn.execute(
                "SELECT * FROM daily_tasks WHERE date=? AND status != 'completed' ORDER BY priority DESC, id",
                (yesterday,)
            ).fetchall()]

    total = len(today_tasks)
    completed = sum(1 for t in today_tasks if t["status"] == "completed")

    return {
        "date": target_date,
        "tasks": today_tasks,
        "rollover": rollover_tasks,
        "summary": {
            "total": total,
            "completed": completed,
            "pending": total - completed,
            "completion_pct": round(completed / total * 100) if total > 0 else 0,
            "rollover_count": len(rollover_tasks),
        }
    }


@router.post("/tasks")
async def create_task(request: TaskRequest):
    """할 일 추가"""
    now = datetime.now().isoformat()
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.execute(
            "INSERT INTO daily_tasks (date, title, priority, status, created_at) VALUES (?, ?, ?, 'pending', ?)",
            (request.date, request.title, request.priority, now)
        )
        tid = cursor.lastrowid
        conn.commit()
    return {"status": "success", "task_id": tid}


@router.patch("/tasks/{task_id}")
async def update_task(task_id: int, request: TaskUpdateRequest):
    """할 일 완료/수정"""
    updates = []
    params = []
    if request.status:
        updates.append("status=?")
        params.append(request.status)
        if request.status == "completed":
            updates.append("completed_at=?")
            params.append(datetime.now().isoformat())
    if request.title:
        updates.append("title=?")
        params.append(request.title)
    if request.priority is not None:
        updates.append("priority=?")
        params.append(request.priority)
    if not updates:
        return {"status": "no_change"}
    params.append(task_id)

    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(f"UPDATE daily_tasks SET {', '.join(updates)} WHERE id=?", tuple(params))
        conn.commit()
    return {"status": "success"}


@router.delete("/tasks/{task_id}")
async def delete_task(task_id: int):
    """할 일 삭제"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM daily_tasks WHERE id=?", (task_id,))
        conn.commit()
    return {"status": "success"}


# ═══════════════════════════════════════════════════════════
# 목표 설정 (일/월/연 — 일일 장부 UI에서 바로 입력)
# ═══════════════════════════════════════════════════════════

class TargetSetRequest(BaseModel):
    target_type: str  # "daily", "monthly", "yearly"
    value: float
    year: Optional[int] = None
    month: Optional[int] = None


@router.get("/targets")
async def get_targets(year: Optional[int] = None):
    """현재 목표 조회 (일일/월별/연간 매출 목표)"""
    if year is None:
        year = date.today().year

    month_key = date.today().strftime("%Y-%m")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row

        # 이번 달 매출 목표
        row = conn.execute(
            "SELECT target_value FROM kpi_targets WHERE month=? AND kpi_key='revenue_krw'",
            (month_key,)
        ).fetchone()
        monthly_target = row["target_value"] if row else 0

        # 올해 전체 매출 목표 (12개월 합)
        row = conn.execute(
            "SELECT COALESCE(SUM(target_value), 0) as total FROM kpi_targets WHERE month LIKE ? AND kpi_key='revenue_krw'",
            (f"{year}-%",)
        ).fetchone()
        yearly_target = row["total"] if row else 0

        # 월별 상세
        monthly_breakdown = [dict(r) for r in conn.execute(
            "SELECT month, target_value FROM kpi_targets WHERE month LIKE ? AND kpi_key='revenue_krw' ORDER BY month",
            (f"{year}-%",)
        ).fetchall()]

    # 하루 목표 = 월 목표 / 이번 달 일수
    import calendar
    today = date.today()
    month_days = calendar.monthrange(today.year, today.month)[1]
    daily_target = round(monthly_target / month_days) if monthly_target > 0 else 0

    return {
        "daily_target": daily_target,
        "monthly_target": monthly_target,
        "yearly_target": yearly_target,
        "year": year,
        "current_month": month_key,
        "monthly_breakdown": monthly_breakdown,
    }


@router.post("/targets")
async def set_target(request: TargetSetRequest):
    """목표 설정 — 일일/월별/연간"""
    today = date.today()
    now = datetime.now().isoformat()

    with sqlite3.connect(DB_PATH) as conn:
        if request.target_type == "monthly":
            # 특정 월의 목표 설정
            month_key = f"{request.year or today.year}-{str(request.month or today.month).zfill(2)}"
            conn.execute("""
                INSERT INTO kpi_targets (month, kpi_key, target_value, unit, updated_at)
                VALUES (?, 'revenue_krw', ?, '원', ?)
                ON CONFLICT(month, kpi_key) DO UPDATE SET
                    target_value=excluded.target_value,
                    updated_at=excluded.updated_at
            """, (month_key, request.value, now))

        elif request.target_type == "yearly":
            # 연 목표를 12개월에 균등 분배
            year = request.year or today.year
            monthly_value = request.value / 12
            for m in range(1, 13):
                month_key = f"{year}-{str(m).zfill(2)}"
                conn.execute("""
                    INSERT INTO kpi_targets (month, kpi_key, target_value, unit, updated_at)
                    VALUES (?, 'revenue_krw', ?, '원', ?)
                    ON CONFLICT(month, kpi_key) DO UPDATE SET
                        target_value=excluded.target_value,
                        updated_at=excluded.updated_at
                """, (month_key, monthly_value, now))

        elif request.target_type == "daily":
            # 일일 목표를 역산해서 월 목표로 저장 (이번 달 기준)
            import calendar
            month_days = calendar.monthrange(today.year, today.month)[1]
            monthly_value = request.value * month_days
            month_key = today.strftime("%Y-%m")
            conn.execute("""
                INSERT INTO kpi_targets (month, kpi_key, target_value, unit, updated_at)
                VALUES (?, 'revenue_krw', ?, '원', ?)
                ON CONFLICT(month, kpi_key) DO UPDATE SET
                    target_value=excluded.target_value,
                    updated_at=excluded.updated_at
            """, (month_key, monthly_value, now))
        else:
            raise HTTPException(400, f"Unknown target_type: {request.target_type}")

        conn.commit()
    return {"status": "success", "target_type": request.target_type}


@router.post("/tasks/{task_id}/rollover")
async def rollover_task(task_id: int, target_date: str):
    """미완료 할 일을 오늘로 이월"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        orig = conn.execute("SELECT * FROM daily_tasks WHERE id=?", (task_id,)).fetchone()
        if not orig:
            raise HTTPException(404, "Task not found")

        now = datetime.now().isoformat()
        conn.execute("""
            INSERT INTO daily_tasks (date, title, priority, status, rollover_from, created_at)
            VALUES (?, ?, ?, 'pending', ?, ?)
        """, (target_date, orig["title"], orig["priority"], orig["date"], now))
        conn.commit()
    return {"status": "success"}


# ═══════════════════════════════════════════════════════════
# 엑셀 업로드 (구매대행 소명자료 장부)
# ═══════════════════════════════════════════════════════════

def _parse_date(val) -> Optional[str]:
    """다양한 형식의 날짜를 ISO 문자열로"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            return datetime.fromisoformat(val.replace(" ", "T")).date().isoformat()
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val, fmt).date().isoformat()
            except Exception:
                continue
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if hasattr(val, 'isoformat'):
        return val.isoformat()[:10]
    return None


def _num(val) -> float:
    """숫자 변환 (실패 시 0)"""
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace(',', '').replace('₩', '').replace('원', '').strip()
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _import_revenue_sheet(ws) -> Dict[str, int]:
    """매출현황 시트 → daily_log 일괄 import"""
    imported, skipped = 0, 0

    # 헤더 위치 확인 (2행이 헤더)
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=col_idx).value
        if v:
            headers[str(v).strip()] = col_idx

    # 필수 컬럼 매핑
    col = {
        "date": headers.get("날짜"),
        "weekday": headers.get("요일"),
        "payment": headers.get("결제금액"),
        "payment_shipping": headers.get("결제배송비"),
        "settlement": headers.get("정산예정금액"),
        "purchase": headers.get("구매금액(원화)"),
        "intl_shipping": headers.get("국제배송비"),
        "cargo_shipping": headers.get("화물택배비"),
        "profit": headers.get("수익금"),
        "orders": headers.get("결제건수"),
        "profit_rate": headers.get("수익률"),
        "fx_usd": headers.get("달러"),
        "fx_cny": headers.get("위안화"),
        "fx_jpy": headers.get("엔화"),
    }

    if not col["date"]:
        return {"imported": 0, "skipped": 0, "error": "날짜 컬럼 없음"}

    now = datetime.now().isoformat()
    with sqlite3.connect(DB_PATH) as conn:
        for row_idx in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=col["date"]).value
            date_str = _parse_date(date_val)
            if not date_str:
                skipped += 1
                continue

            payment = _num(ws.cell(row=row_idx, column=col["payment"]).value) if col["payment"] else 0

            # 결제금액 0인 행은 스킵 (환율만 있는 행 등)
            if payment == 0:
                skipped += 1
                continue

            weekday = ws.cell(row=row_idx, column=col["weekday"]).value if col["weekday"] else ''
            payment_shipping = _num(ws.cell(row=row_idx, column=col["payment_shipping"]).value) if col["payment_shipping"] else 0
            settlement = _num(ws.cell(row=row_idx, column=col["settlement"]).value) if col["settlement"] else 0
            purchase = _num(ws.cell(row=row_idx, column=col["purchase"]).value) if col["purchase"] else 0
            intl_shipping = _num(ws.cell(row=row_idx, column=col["intl_shipping"]).value) if col["intl_shipping"] else 0
            cargo_shipping = _num(ws.cell(row=row_idx, column=col["cargo_shipping"]).value) if col["cargo_shipping"] else 0
            profit = _num(ws.cell(row=row_idx, column=col["profit"]).value) if col["profit"] else 0
            orders = int(_num(ws.cell(row=row_idx, column=col["orders"]).value)) if col["orders"] else 0
            profit_rate = _num(ws.cell(row=row_idx, column=col["profit_rate"]).value) if col["profit_rate"] else 0
            fx_usd = _num(ws.cell(row=row_idx, column=col["fx_usd"]).value) if col["fx_usd"] else 0
            fx_cny = _num(ws.cell(row=row_idx, column=col["fx_cny"]).value) if col["fx_cny"] else 0
            fx_jpy = _num(ws.cell(row=row_idx, column=col["fx_jpy"]).value) if col["fx_jpy"] else 0

            conn.execute("""
                INSERT INTO daily_log (
                    date, weekday, payment_krw, payment_shipping_krw, settlement_krw,
                    purchase_cost_krw, international_shipping_krw, cargo_shipping_krw,
                    profit_krw, orders_count, profit_rate,
                    fx_usd, fx_cny, fx_jpy,
                    revenue_krw, cost_krw, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(date) DO UPDATE SET
                    weekday=excluded.weekday,
                    payment_krw=excluded.payment_krw,
                    payment_shipping_krw=excluded.payment_shipping_krw,
                    settlement_krw=excluded.settlement_krw,
                    purchase_cost_krw=excluded.purchase_cost_krw,
                    international_shipping_krw=excluded.international_shipping_krw,
                    cargo_shipping_krw=excluded.cargo_shipping_krw,
                    profit_krw=excluded.profit_krw,
                    orders_count=excluded.orders_count,
                    profit_rate=excluded.profit_rate,
                    fx_usd=excluded.fx_usd, fx_cny=excluded.fx_cny, fx_jpy=excluded.fx_jpy,
                    revenue_krw=excluded.revenue_krw, cost_krw=excluded.cost_krw,
                    updated_at=excluded.updated_at
            """, (
                date_str, str(weekday or ''), payment, payment_shipping, settlement,
                purchase, intl_shipping, cargo_shipping,
                profit, orders, profit_rate,
                fx_usd, fx_cny, fx_jpy,
                payment,  # revenue_krw = payment_krw (하위 호환)
                purchase + intl_shipping + cargo_shipping,  # cost_krw
                now
            ))
            imported += 1
        conn.commit()

    return {"imported": imported, "skipped": skipped}


def _import_order_sheet(ws) -> Dict[str, int]:
    """주문현황 시트 → order_ledger"""
    imported, skipped = 0, 0
    now = datetime.now().isoformat()

    # 헤더 매핑 (1행)
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v:
            headers[str(v).strip()] = col_idx

    def g(row, key):
        idx = headers.get(key)
        return ws.cell(row=row, column=idx).value if idx else None

    with sqlite3.connect(DB_PATH) as conn:
        for row_idx in range(2, ws.max_row + 1):
            order_no = g(row_idx, "주문번호")
            if not order_no:
                skipped += 1
                continue

            order_date = _parse_date(g(row_idx, "주문일"))
            if not order_date:
                skipped += 1
                continue

            try:
                conn.execute("""
                    INSERT INTO order_ledger (
                        order_date, platform, order_no, product_no, product_name, option_name,
                        quantity, product_url, buyer_name, buyer_phone, recipient_name, recipient_phone,
                        customs_no, zipcode, address, delivery_memo,
                        payment_krw, shipping_krw, settlement_krw,
                        order_placed_date, source_vendor, source_order_no, source_tracking,
                        payment_card, foreign_currency, foreign_amount, purchase_usd, purchase_krw,
                        forwarding_center, domestic_carrier, domestic_tracking, international_shipping_krw,
                        cargo_carrier, cargo_tracking, cargo_shipping_krw, customs_duty_krw,
                        profit_krw, profit_rate, note, imported_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(order_no, platform) DO UPDATE SET
                        product_name=excluded.product_name,
                        payment_krw=excluded.payment_krw,
                        settlement_krw=excluded.settlement_krw,
                        purchase_krw=excluded.purchase_krw,
                        profit_krw=excluded.profit_krw,
                        imported_at=excluded.imported_at
                """, (
                    order_date, str(g(row_idx, "주문플랫폼") or ''), str(order_no),
                    str(g(row_idx, "상품번호") or ''), str(g(row_idx, "상품명") or ''),
                    str(g(row_idx, "옵션") or ''), int(_num(g(row_idx, "수량"))) or 1,
                    str(g(row_idx, "상품URL") or ''), str(g(row_idx, "구매자명") or ''),
                    str(g(row_idx, "구매자연락처") or ''), str(g(row_idx, "수취인명") or ''),
                    str(g(row_idx, "수취인연락처") or ''), str(g(row_idx, "통관번호") or ''),
                    str(g(row_idx, "우편번호") or ''), str(g(row_idx, "주소") or ''),
                    str(g(row_idx, "배송메세지") or ''),
                    _num(g(row_idx, "상품결제금액")), _num(g(row_idx, "결제배송비")), _num(g(row_idx, "정산예정금액")),
                    _parse_date(g(row_idx, "발주일자")) or '',
                    str(g(row_idx, "해외구매처") or ''), str(g(row_idx, "해외주문번호") or ''),
                    str(g(row_idx, "해외송장번호") or ''), str(g(row_idx, "결제카드") or ''),
                    str(g(row_idx, "해외현지화폐") or ''), _num(g(row_idx, "해외결제금액")),
                    _num(g(row_idx, "구매금액(USD)")), _num(g(row_idx, "구매금액(원화)")),
                    str(g(row_idx, "배송대행지") or ''), str(g(row_idx, "국내택배사") or ''),
                    str(g(row_idx, "국내운송장번호") or ''), _num(g(row_idx, "국제배송비")),
                    str(g(row_idx, "화물택배사") or ''), str(g(row_idx, "화물운송장번호") or ''),
                    _num(g(row_idx, "화물택배비")), _num(g(row_idx, "관부가세")),
                    _num(g(row_idx, "수익금")), _num(g(row_idx, "수익률")),
                    str(g(row_idx, "비고") or ''), now
                ))
                imported += 1
            except Exception as e:
                logger.warning(f"주문 import 실패 row {row_idx}: {e}")
                skipped += 1
        conn.commit()

    return {"imported": imported, "skipped": skipped}


def _import_cs_templates(ws) -> Dict[str, int]:
    """CS 템플릿 시트 → cs_templates"""
    imported = 0
    now = datetime.now().isoformat()
    with sqlite3.connect(DB_PATH) as conn:
        for row_idx in range(1, ws.max_row + 1):
            category = ws.cell(row=row_idx, column=1).value
            template = ws.cell(row=row_idx, column=2).value
            if category and template:
                conn.execute("""
                    INSERT INTO cs_templates (category, template, source, created_at)
                    VALUES (?, ?, 'excel_import', ?)
                    ON CONFLICT(category) DO UPDATE SET
                        template=excluded.template,
                        created_at=excluded.created_at
                """, (str(category).strip(), str(template).strip(), now))
                imported += 1
        conn.commit()
    return {"imported": imported}


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """엑셀 업로드 — 구매대행 소명자료 장부 일괄 import"""
    try:
        import openpyxl
        from io import BytesIO

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

        results = {
            "filename": file.filename,
            "sheets_found": wb.sheetnames,
            "revenue": None,
            "orders": None,
            "cs_templates": None,
        }

        # 매출현황 시트 → daily_log
        if "매출현황" in wb.sheetnames:
            results["revenue"] = _import_revenue_sheet(wb["매출현황"])

        # 주문현황 시트 → order_ledger
        if "주문현황" in wb.sheetnames:
            results["orders"] = _import_order_sheet(wb["주문현황"])

        # CS 템플릿 → cs_templates
        if "CS 템플릿" in wb.sheetnames:
            results["cs_templates"] = _import_cs_templates(wb["CS 템플릿"])

        wb.close()
        return {"status": "success", "results": results}

    except Exception as e:
        logger.exception("엑셀 import 실패")
        return {"status": "failed", "error": str(e)}


@router.get("/orders/recent")
async def recent_orders(limit: int = 20):
    """최근 주문 목록 (장부)"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM order_ledger ORDER BY order_date DESC, id DESC LIMIT ?",
            (limit,)
        ).fetchall()
    return {"orders": [dict(r) for r in rows], "total": len(rows)}


@router.get("/cs-templates")
async def list_cs_templates():
    """CS 템플릿 목록"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM cs_templates ORDER BY id").fetchall()
    return {"templates": [dict(r) for r in rows]}


# ═══════════════════════════════════════════════════════════
# 세무 증빙 장부 — 거래 원장 / 월별 집계 / CSV 내보내기
# ═══════════════════════════════════════════════════════════

@router.get("/ledger/orders")
async def ledger_orders(
    year_month: Optional[str] = None,  # "2025-05"
    platform: Optional[str] = None,
    q: Optional[str] = None,           # 검색어 (상품명/구매자/통관번호)
    limit: int = 500,
    offset: int = 0,
):
    """거래 원장 — 필터링 가능한 주문 목록"""
    sql = "SELECT * FROM order_ledger WHERE 1=1"
    params: List[Any] = []
    if year_month:
        sql += " AND tax_year_month = ?"
        params.append(year_month)
    if platform:
        sql += " AND platform = ?"
        params.append(platform)
    if q:
        sql += " AND (product_name LIKE ? OR buyer_name LIKE ? OR recipient_name LIKE ? OR customs_no LIKE ? OR order_no LIKE ?)"
        kw = f"%{q}%"
        params.extend([kw, kw, kw, kw, kw])
    sql += " ORDER BY order_date DESC, id DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql, params).fetchall()
        # 전체 카운트 (페이징용)
        cnt_sql = "SELECT COUNT(*) FROM order_ledger WHERE 1=1"
        cnt_params: List[Any] = []
        if year_month:
            cnt_sql += " AND tax_year_month = ?"
            cnt_params.append(year_month)
        if platform:
            cnt_sql += " AND platform = ?"
            cnt_params.append(platform)
        if q:
            cnt_sql += " AND (product_name LIKE ? OR buyer_name LIKE ? OR recipient_name LIKE ? OR customs_no LIKE ? OR order_no LIKE ?)"
            kw = f"%{q}%"
            cnt_params.extend([kw, kw, kw, kw, kw])
        total = conn.execute(cnt_sql, cnt_params).fetchone()[0]

    return {
        "orders": [dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.get("/ledger/months")
async def ledger_months():
    """거래가 있는 월 목록 (드롭다운용)"""
    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute("""
            SELECT tax_year_month as ym, COUNT(*) as cnt
            FROM order_ledger
            WHERE tax_year_month != ''
            GROUP BY tax_year_month
            ORDER BY tax_year_month DESC
        """).fetchall()
    return {"months": [{"ym": r[0], "count": r[1]} for r in rows]}


@router.get("/tax/monthly")
async def tax_monthly(year_month: str):
    """월별 세무 집계 — 부가세 신고용"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        # 전체 합계
        total = conn.execute("""
            SELECT
                COUNT(*) as orders,
                COALESCE(SUM(payment_krw), 0) as revenue,
                COALESCE(SUM(settlement_krw), 0) as settlement,
                COALESCE(SUM(platform_fee_krw), 0) as platform_fee,
                COALESCE(SUM(purchase_krw), 0) as purchase,
                COALESCE(SUM(international_shipping_krw), 0) as intl_ship,
                COALESCE(SUM(cargo_shipping_krw), 0) as cargo_ship,
                COALESCE(SUM(customs_duty_krw), 0) as customs_duty,
                COALESCE(SUM(profit_krw), 0) as profit,
                COALESCE(SUM(vat_krw), 0) as vat
            FROM order_ledger
            WHERE tax_year_month = ?
        """, (year_month,)).fetchone()

        # 플랫폼별 내역
        by_platform = conn.execute("""
            SELECT
                platform,
                COUNT(*) as orders,
                COALESCE(SUM(payment_krw), 0) as revenue,
                COALESCE(SUM(platform_fee_krw), 0) as fee,
                COALESCE(SUM(profit_krw), 0) as profit
            FROM order_ledger
            WHERE tax_year_month = ?
            GROUP BY platform
            ORDER BY revenue DESC
        """, (year_month,)).fetchall()

        # 과세구분별 (설정된 건만)
        by_tax = conn.execute("""
            SELECT
                CASE WHEN tax_category = '' OR tax_category IS NULL THEN '미분류' ELSE tax_category END as cat,
                COUNT(*) as orders,
                COALESCE(SUM(payment_krw), 0) as revenue
            FROM order_ledger
            WHERE tax_year_month = ?
            GROUP BY cat
        """, (year_month,)).fetchall()

    d = dict(total) if total else {}
    # 부가세 추정 (수수료 매출 기준 가정 — 수익금 * 10/110)
    estimated_vat_on_profit = round(d.get("profit", 0) * 10 / 110)
    # 부가세 추정 (전체 매출 기준 — 위탁매매 가정)
    estimated_vat_on_revenue = round(d.get("revenue", 0) * 10 / 110)

    return {
        "year_month": year_month,
        "total": d,
        "by_platform": [dict(r) for r in by_platform],
        "by_tax_category": [dict(r) for r in by_tax],
        "vat_estimates": {
            "on_profit_only": estimated_vat_on_profit,  # 단순대행(수수료만 과세)
            "on_full_revenue": estimated_vat_on_revenue,  # 위탁매매(전체 과세)
            "note": "[확인 필요] 세무사와 과세구조 확정 후 하나 선택",
        },
    }


@router.get("/tax/export")
async def tax_export(year_month: str):
    """월별 거래원장 CSV 내보내기 — 세무사 제출용"""
    import csv
    from io import StringIO
    from fastapi.responses import Response

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT
                order_date, platform, order_no, product_name, option_name, quantity,
                buyer_name, recipient_name, customs_no, zipcode, address,
                payment_krw, shipping_krw, settlement_krw, platform_fee_krw,
                source_vendor, foreign_currency, foreign_amount, purchase_krw,
                international_shipping_krw, cargo_shipping_krw, customs_duty_krw,
                profit_krw, profit_rate, tax_category, receipt_type, receipt_no, note
            FROM order_ledger
            WHERE tax_year_month = ?
            ORDER BY order_date, id
        """, (year_month,)).fetchall()

    out = StringIO()
    out.write('\ufeff')  # UTF-8 BOM (엑셀 한글 호환)
    w = csv.writer(out)
    w.writerow([
        "주문일", "플랫폼", "주문번호", "상품명", "옵션", "수량",
        "구매자", "수취인", "통관번호", "우편번호", "주소",
        "결제금액", "결제배송비", "정산예정금액", "플랫폼수수료",
        "해외구매처", "해외통화", "해외결제금액", "구매원가(원)",
        "국제배송비", "화물택배비", "관부가세",
        "수익금", "수익률", "과세구분", "증빙종류", "증빙번호", "비고",
    ])
    for r in rows:
        w.writerow([r[k] if r[k] is not None else '' for k in r.keys()])

    csv_bytes = out.getvalue().encode('utf-8')
    filename = f"fortimove_ledger_{year_month}.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class OrderTaxUpdate(BaseModel):
    tax_category: Optional[str] = None
    platform_fee_krw: Optional[float] = None
    pg_fee_krw: Optional[float] = None
    vat_krw: Optional[float] = None
    receipt_type: Optional[str] = None
    receipt_no: Optional[str] = None
    card_last4: Optional[str] = None
    bank_tx_id: Optional[str] = None
    note: Optional[str] = None


@router.patch("/ledger/order/{order_id}")
async def update_order_tax(order_id: int, payload: OrderTaxUpdate):
    """개별 주문의 세무 필드 수정 (과세구분/수수료/증빙번호 등)"""
    fields = {k: v for k, v in payload.dict().items() if v is not None}
    if not fields:
        return {"status": "noop"}
    cols = ", ".join(f"{k} = ?" for k in fields)
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            f"UPDATE order_ledger SET {cols} WHERE id = ?",
            list(fields.values()) + [order_id],
        )
        conn.commit()
    return {"status": "updated", "order_id": order_id, "fields": list(fields.keys())}
